import os
import time
import urllib.request
import urllib.parse

from openpyxl import load_workbook

failed_counter = 0
undeleted_files = 0
success_counter = 0


# Check if folder exists, if not create folder
def check_folder_exists(folder_name):
    if not os.path.isdir(folder_name):
        os.makedirs(folder_name)


def check_file_exists(file_name):
    return os.path.isfile(file_name)


# def convert_doc2txt(source_path, destination_path):
#     global failed_counter
#     global success_counter
#     global undeleted_files
#     print("convert: ", source_path)
#     # Passing docx file to process function
#     try:
#         doc = aw.Document(source_path)
#         doc.save(destination_path)
#     except:
#         failed_counter += 1
#         try:
#             os.remove(source_path)
#         except:
#             undeleted_files += 1
#         return 0
#     success_counter += 1
#     # Saving content inside docx file into output.txt file
#     os.remove(source_path)
#     return 1


def download_from_url_with_repeats(url, path):
    f = open("log_error.txt", "a")
    remaining_download_tries = 15
    while remaining_download_tries > 0:
        try:
            urllib.request.urlretrieve(url, path)
            print("successfully downloaded: " + path)
        except:
            print("error downloading from " + url + " on trial no: " + str(16 - remaining_download_tries))
            remaining_download_tries = remaining_download_tries - 1
            time.sleep(16)
            continue
        else:
            break
    if remaining_download_tries == 0:
        f.write("error downloading " + url + " to " + path + "\n")


def download_bills(knesset):
    xlsx_path = "{}.xlsx".format(knesset)
    wb = load_workbook(xlsx_path)
    ws = wb.active
    for row in ws.iter_rows():
        # download file
        if row[0].value is None:
            continue
        gender = "{}".format(row[4].value)
        url = row[5].value
        if gender == "זכר":
            gender_name = "m"
        else:
            gender_name = "f"
        download_file_path = "downloadedFiles"
        check_folder_exists(download_file_path)
        file_name = "/{}_{}_{}".format(knesset, gender_name, row[0].value)
        if check_file_exists(download_file_path + file_name + ".doc"):
            continue
        url = url.replace(" ", "%20")
        download_from_url_with_repeats(url, download_file_path + file_name + ".doc")

        # convert to txt
        # status = convert_doc2txt(download_file_path + file_name + ".doc", download_file_path + file_name + ".txt")
        # # # remove doc file
        # # os.remove(download_file_path + file_name + ".doc")
        # # extract information
        # if status:
        #     trim_file(download_file_path + file_name + ".txt")


def init_downloads():
    knesset_numbers = ["K17", "K18", "K20"]
    for knesset in knesset_numbers:
        download_bills(knesset)
        print(knesset + "is Done")


init_downloads()
print(success_counter, failed_counter, undeleted_files)
