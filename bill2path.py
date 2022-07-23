from openpyxl import load_workbook
import csv


def get_filename_and_path(knesset_num):
    global dict_data
    xlsx_path = "{}.xlsx".format(knesset_num)
    wb = load_workbook(xlsx_path)
    ws = wb.active
    for row in ws.iter_rows():
        # download file
        if row[0].value is None:
            continue
        gender = "{}".format(row[4].value)
        url = row[5].value
        url = url.replace(" ", "%20")
        if gender == "זכר":
            gender_name = "m"
        else:
            gender_name = "f"
        file_name = "{}_{}_{}".format(knesset_num, gender_name, row[0].value)
        bill_dic = {file_name: url}
        dict_data = dict(dict_data, **bill_dic)


knesset_numbers = ["K17", "K18", "K20"]
k17_topics = ["מינוי בכירים",
              "שיכון ובינוי",
              "זכויות עובדים",
              "פלילים",
              "בנקאות",
              "אלימות במשפחה",
              "זכויות נשים",
              "צרכנות",
              "קצבאות ותגמולים",
              "דת ומדינה",
              "תקשורת",
              "זכויות אדם",
              "תכנון ובנייה",
              "משפחה וילדים",
              "תחבורה",
              "בריאות",
              "שיוויון מגדרי",
              "איכות הסביבה",
              "אנרגיה",
              "בחירות",
              "בטחון וצבא",
              "עלייה וקליטה",
              "חינוך",
              "דיור ונדלן"]
k17_merge_topics = [(3, 24), (18, 29)]
k17_redundant_topics = [5, 6, 14, 27]
k18_topics = ["תקשורת",
              "מדיניות כלכלית",
              "בנקאות והשקעות",
              "איכות הסביבה",
              "מיסוי ושכר",
              "בטחון פנים",
              "פלילים",
              "תחבורה",
              "שירות צבאי",
              "מעמד אזרחי",
              "פיתוח אזורי",
              "בחירות",
              "זכויות אדם",
              "עבירות מין",
              "תכנון ובנייה",
              "חינוך",
              "מינוי בכירים",
              "מורשת יהודית",
              "דת ומדינה",
              "משפט",
              "קצבאות ותגמולים",
              "זכויות עובדים",
              "דיור ונדלן",
              "בריאות",
              "צרכנות"]
k18_merge_topics = [(4, 18)]
k18_redundant_topics = [6, 9, 16, 22]
k20_topics = ["תכנון ובנייה",
              "תחבורה",
              "משפחה וילדים",
              "שירות צבאי",
              "פלילים",
              "חינוך",
              "בריאות",
              "רווחה",
              "קצבאות ותגמולים",
              "מדיניות כלכלית",
              "תרבות וספורט",
              "זכויות עובדים",
              "זכויות אדם",
              "דיור ונדלן",
              "איכות הסביבה",
              "בנקאות",
              "תעסוקה",
              "צרכנות",
              "תקשורת",
              "בחירות",
              "זכויות נשים",
              "משפט",
              "דת ומדינה"]
k20_merge_topics = [(5, 26), (15, 19), (17, 20)]
k20_redundant_topics = [4, 10, 13, 18]
num_of_topics = 30
csv_columns = ["FileName", "Path"]
dict_data = {}
K17_m_topic_array = []
for i in range(30):
    K17_m_topic_array.append([])
K17_f_topic_array = []
for i in range(30):
    K17_f_topic_array.append([])
K18_m_topic_array = []
for i in range(30):
    K18_m_topic_array.append([])
K18_f_topic_array = []
for i in range(30):
    K18_f_topic_array.append([])
K20_m_topic_array = []
for i in range(30):
    K20_m_topic_array.append([])
K20_f_topic_array = []
for i in range(30):
    K20_f_topic_array.append([])

for knesset in knesset_numbers:
    get_filename_and_path(knesset)

for knesset in knesset_numbers:
    path = "./{}_document-topic-distribution.csv".format(knesset)
    with open(path, newline='', encoding='utf-8') as csvfile:
        spamreader = csv.reader(csvfile, delimiter=' ', quotechar='|')
        i = 0
        for row in spamreader:
            string = row[0]
            split = string.split(";")
            if i == 0:
                i += 1
                continue
            max_value = max(split[1:])
            max_index = split.index(max_value) - 1
            bill_path = dict_data.get(split[0])
            split_file_name = split[0].split("_")
            if split_file_name[0] == "K17":
                if split_file_name[1] == "m":
                    (K17_m_topic_array[max_index]).append(bill_path)
                else:
                    (K17_f_topic_array[max_index]).append(bill_path)
            elif split_file_name[0] == "K18":
                if split_file_name[1] == "m":
                    K18_m_topic_array[max_index].append(bill_path)
                else:
                    K18_f_topic_array[max_index].append(bill_path)
            else:
                if split_file_name[1] == "m":
                    K20_m_topic_array[max_index].append(bill_path)
                else:
                    K20_f_topic_array[max_index].append(bill_path)

with open("K17_m_topic2path.csv", "w", encoding='UTF8') as f:
    topic_to_remove = []
    for (a, b) in k17_merge_topics:
        K17_m_topic_array[a] += K17_m_topic_array[b]
        topic_to_remove.append(b)
    for topic in k17_redundant_topics:
        topic_to_remove.append(topic)
    topic_to_remove.sort(reverse=True)
    for topic in topic_to_remove:
        K17_m_topic_array.pop(topic)
    for i in range(len(k17_topics)):
        K17_m_topic_array[i].insert(0, k17_topics[i])
    writer = csv.writer(f)
    writer.writerows(K17_m_topic_array)

with open("K17_f_topic2path.csv", "w", encoding='UTF8') as f:
    topic_to_remove = []
    for (a, b) in k17_merge_topics:
        K17_f_topic_array[a] += K17_f_topic_array[b]
        topic_to_remove.append(b)
    for topic in k17_redundant_topics:
        topic_to_remove.append(topic)
    topic_to_remove.sort(reverse=True)
    for topic in topic_to_remove:
        K17_f_topic_array.pop(topic)
    for i in range(len(k17_topics)):
        K17_f_topic_array[i].insert(0, k17_topics[i])
    writer = csv.writer(f)
    writer.writerows(K17_f_topic_array)

with open("K18_m_topic2path.csv", "w", encoding='UTF8') as f:
    topic_to_remove = []
    for (a, b) in k18_merge_topics:
        K18_m_topic_array[a] += K18_m_topic_array[b]
        topic_to_remove.append(b)
    for topic in k18_redundant_topics:
        topic_to_remove.append(topic)
    topic_to_remove.sort(reverse=True)
    for topic in topic_to_remove:
        K18_m_topic_array.pop(topic)
    for i in range(len(k18_topics)):
        K18_m_topic_array[i].insert(0, k18_topics[i])
    writer = csv.writer(f)
    writer.writerows(K18_m_topic_array)

with open("K18_f_topic2path.csv", "w", encoding='UTF8') as f:
    topic_to_remove = []
    for (a, b) in k18_merge_topics:
        K18_f_topic_array[a] += K18_f_topic_array[b]
        topic_to_remove.append(b)
    for topic in k18_redundant_topics:
        topic_to_remove.append(topic)
    topic_to_remove.sort(reverse=True)
    for topic in topic_to_remove:
        K18_f_topic_array.pop(topic)
    for i in range(len(k18_topics)):
        K18_f_topic_array[i].insert(0, k18_topics[i])
    writer = csv.writer(f)
    writer.writerows(K18_f_topic_array)

with open("K20_m_topic2path.csv", "w", encoding='UTF8') as f:
    topic_to_remove = []
    for (a, b) in k20_merge_topics:
        K20_m_topic_array[a] += K20_m_topic_array[b]
        topic_to_remove.append(b)
    for topic in k20_redundant_topics:
        topic_to_remove.append(topic)
    topic_to_remove.sort(reverse=True)
    for topic in topic_to_remove:
        K20_m_topic_array.pop(topic)
    for i in range(len(k20_topics)):
        K20_m_topic_array[i].insert(0, k20_topics[i])
    writer = csv.writer(f)
    writer.writerows(K20_m_topic_array)

with open("K20_f_topic2path.csv", "w", encoding='UTF8') as f:
    topic_to_remove = []
    for (a, b) in k20_merge_topics:
        K20_f_topic_array[a] += K20_f_topic_array[b]
        topic_to_remove.append(b)
    for topic in k20_redundant_topics:
        topic_to_remove.append(topic)
    topic_to_remove.sort(reverse=True)
    for topic in topic_to_remove:
        K20_f_topic_array.pop(topic)
    for i in range(len(k20_topics)):
        K20_f_topic_array[i].insert(0, k20_topics[i])
    writer = csv.writer(f)
    writer.writerows(K20_f_topic_array)
